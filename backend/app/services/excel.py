"""Đọc/ghi file Excel cho chức năng nhập học sinh."""

import io

import openpyxl

from app.schemas.people import StudentImportRow

# Các tên cột có thể gặp (đã chuẩn hóa lower)
NAME_COLS = {"họ tên", "ho ten", "họ và tên", "ho va ten", "name", "fullname", "tên"}
EMAIL_COLS = {"email", "e-mail", "thư điện tử"}
CLASS_COLS = {"lớp", "lop", "class", "classname", "tên lớp"}
STATUS_COLS = {"trạng thái", "trang thai", "status"}


def _find_index(headers: list[str], candidates: set[str]) -> int | None:
    for i, h in enumerate(headers):
        if h in candidates:
            return i
    return None


def parse_students(content: bytes) -> list[StudentImportRow]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    i_name = _find_index(headers, NAME_COLS)
    i_email = _find_index(headers, EMAIL_COLS)
    i_class = _find_index(headers, CLASS_COLS)
    i_status = _find_index(headers, STATUS_COLS)

    def cell(row: tuple, idx: int | None) -> str:
        if idx is None or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    out: list[StudentImportRow] = []
    for row in rows[1:]:
        name = cell(row, i_name)
        if not name:
            continue
        out.append(
            StudentImportRow(
                name=name,
                email=cell(row, i_email) or None,
                className=cell(row, i_class) or None,
                status=cell(row, i_status) or None,
            )
        )
    return out


def build_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HocSinh"
    ws.append(["Họ tên", "Email", "Lớp", "Trạng thái"])
    ws.append(["Nguyễn Văn A", "a@hs.cva.edu.vn", "Lớp 10A1", "Đang học"])
    ws.append(["Trần Thị B", "b@hs.cva.edu.vn", "Lớp 10A2", "Đang học"])
    for col, width in zip("ABCD", (22, 26, 12, 12)):
        ws.column_dimensions[col].width = width
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
